#!/usr/bin/env python3
"""
Find fields whose name contains 'date' or 'timestamp' but whose
data type is a string type (varchar, char, text, tinytext, mediumtext,
longtext, enum, set) on either source or target.

Output: date_as_string.xlsx
"""

import json, os, sys
import MySQLdb
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
with open(CONFIG_PATH) as f:
    cfg = json.load(f)

SRC = cfg["source"]
TGT = cfg["target"]
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from mysql40 import MySQL40Connection

SYS_DBS = {"information_schema", "performance_schema", "mysql", "sys"}
STRING_TYPES = ("varchar", "char", "text", "tinytext", "mediumtext", "longtext", "enum", "set")
DATE_KEYWORDS = ("date", "timestamp")


def is_string_type(t):
    t = t.lower()
    return any(t.startswith(s) for s in STRING_TYPES)


def has_date_keyword(name):
    n = name.lower()
    return any(kw in n for kw in DATE_KEYWORDS)


def get_databases():
    conn = MySQLdb.connect(host=TGT["host"], port=int(TGT["port"]),
                           user=TGT["user"], passwd=TGT["password"], connect_timeout=10)
    cur = conn.cursor()
    cur.execute("SHOW DATABASES")
    dbs = [r[0] for r in cur.fetchall() if r[0].lower() not in SYS_DBS]
    cur.close(); conn.close()
    return dbs


def get_schema_src(db):
    conn = MySQL40Connection(host=SRC["host"], port=int(SRC["port"]),
                             user=SRC["user"], password=SRC["password"],
                             database=db, timeout=60, charset="tis620")
    schema = {}
    for row in conn.query("SHOW TABLES"):
        tbl = row[0]
        schema[tbl] = {}
        for col in conn.query(f"SHOW COLUMNS FROM `{tbl}`"):
            schema[tbl][col[0]] = col[1]  # field -> type
    return schema


def get_schema_tgt(db):
    conn = MySQLdb.connect(host=TGT["host"], port=int(TGT["port"]),
                           user=TGT["user"], passwd=TGT["password"],
                           db=db, charset="utf8mb4", connect_timeout=10)
    cur = conn.cursor()
    cur.execute("SHOW TABLES")
    tables = [r[0] for r in cur.fetchall()]
    schema = {}
    for tbl in tables:
        cur.execute(f"SHOW COLUMNS FROM `{tbl}`")
        schema[tbl] = {col[0]: col[1] for col in cur.fetchall()}
    cur.close(); conn.close()
    return schema


# ── build rows ─────────────────────────────────────────────────────────────────
all_rows = []  # (db, table, field, src_type, tgt_type, flag)

databases = get_databases()
print(f"Databases: {databases}\n")

for db in databases:
    print(f"[{db}]", end=" ", flush=True)
    try:
        src = get_schema_src(db)
    except Exception as e:
        print(f"src ERR: {e}", end=" ")
        src = {}
    try:
        tgt = get_schema_tgt(db)
    except Exception as e:
        print(f"tgt ERR: {e}", end=" ")
        tgt = {}

    all_tables = sorted(set(list(src.keys()) + list(tgt.keys())))
    count = 0
    for tbl in all_tables:
        sf = src.get(tbl, {})
        tf = tgt.get(tbl, {})
        all_fields = list(dict.fromkeys(list(sf.keys()) + list(tf.keys())))
        for field in all_fields:
            if not has_date_keyword(field):
                continue
            src_type = sf.get(field, "—")
            tgt_type = tf.get(field, "—")
            src_str = is_string_type(src_type) if src_type != "—" else False
            tgt_str = is_string_type(tgt_type) if tgt_type != "—" else False
            if not src_str and not tgt_str:
                continue
            flag = []
            if src_str: flag.append("source")
            if tgt_str: flag.append("target")
            all_rows.append((db, tbl, field, src_type, tgt_type, ", ".join(flag)))
            count += 1
    print(f"{count} found")

# ── write Excel ────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "date_as_string"

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
SRC_FILL = PatternFill("solid", fgColor="FCE4D6")  # orange-red = string on source
TGT_FILL = PatternFill("solid", fgColor="FFF2CC")  # yellow = string on target
BOTH_FILL = PatternFill("solid", fgColor="F4CCCC") # red = both

headers = ["Database", "Table", "Field", "Source Type (MySQL4)", "Target Type (MySQL8)", "String On"]
ws.append(headers)
for ci, h in enumerate(headers, 1):
    cell = ws.cell(1, ci)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")

for db, tbl, field, src_type, tgt_type, flag in all_rows:
    ws.append([db, tbl, field, src_type, tgt_type, flag])
    row = ws[ws.max_row]
    if "source" in flag and "target" in flag:
        fill = BOTH_FILL
    elif "source" in flag:
        fill = SRC_FILL
    else:
        fill = TGT_FILL
    for cell in row:
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")

col_widths = [12, 35, 30, 26, 26, 14]
for ci, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:F{ws.max_row}"

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "date_as_string.xlsx")
wb.save(out)
print(f"\nTotal: {len(all_rows)} fields → {out}")
