#!/usr/bin/env python3
"""
Dump schema (fields, data types, PK) for every table in every database
from both source (MySQL 4) and target (MySQL 8) into one Excel file.

Output: schema_dump.xlsx
  - One sheet per database
  - Columns: Database | Table | Field | Source Type | Target Type | Is PK | PK Order
"""

import json
import os
import sys

import MySQLdb
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── load config ────────────────────────────────────────────────────────────────
CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
with open(CONFIG_PATH, "r", encoding="utf-8") as f:
    cfg = json.load(f)

SRC = cfg["source"]
TGT = cfg["target"]

# ── MySQL 4 connection ─────────────────────────────────────────────────────────
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from mysql40 import MySQL40Connection

SYS_DBS = {"information_schema", "performance_schema", "mysql", "sys"}

# ── helpers ────────────────────────────────────────────────────────────────────

def get_databases_tgt():
    conn = MySQLdb.connect(
        host=TGT["host"], port=int(TGT["port"]),
        user=TGT["user"], passwd=TGT["password"],
        connect_timeout=10,
    )
    cur = conn.cursor()
    cur.execute("SHOW DATABASES")
    dbs = [r[0] for r in cur.fetchall() if r[0].lower() not in SYS_DBS]
    cur.close(); conn.close()
    return dbs


def get_schema_src(db):
    """Returns {table: {field: {type, pk_order}}} from MySQL 4 source."""
    conn = MySQL40Connection(
        host=SRC["host"], port=int(SRC["port"]),
        user=SRC["user"], password=SRC["password"],
        database=db, timeout=60, charset="tis620",
    )
    tables_raw = conn.query("SHOW TABLES")
    schema = {}
    for row in tables_raw:
        tbl = row[0]
        cols_raw = conn.query(f"SHOW COLUMNS FROM `{tbl}`")
        # SHOW COLUMNS: Field, Type, Null, Key, Default, Extra
        fields = {}
        pk_order = 0
        for col in cols_raw:
            fname, ftype, _, key, _, _ = col[0], col[1], col[2], col[3], col[4], col[5]
            is_pk = (key == "PRI")
            if is_pk:
                pk_order += 1
            fields[fname] = {"type": ftype, "pk_order": pk_order if is_pk else 0}
        schema[tbl] = fields
    return schema


def get_schema_tgt(db):
    """Returns {table: {field: {type, pk_order}}} from MySQL 8 target."""
    conn = MySQLdb.connect(
        host=TGT["host"], port=int(TGT["port"]),
        user=TGT["user"], passwd=TGT["password"],
        db=db, charset="utf8mb4", connect_timeout=10,
    )
    cur = conn.cursor()
    cur.execute("SHOW TABLES")
    tables = [r[0] for r in cur.fetchall()]
    schema = {}
    for tbl in tables:
        cur.execute(f"SHOW COLUMNS FROM `{tbl}`")
        fields = {}
        pk_order = 0
        for col in cur.fetchall():
            fname, ftype, _, key = col[0], col[1], col[2], col[3]
            is_pk = (key == "PRI")
            if is_pk:
                pk_order += 1
            fields[fname] = {"type": ftype, "pk_order": pk_order if is_pk else 0}
        schema[tbl] = fields
    cur.close(); conn.close()
    return schema


STRING_TYPES  = ("varchar", "char", "text", "tinytext", "mediumtext", "longtext", "enum", "set")
DATE_KEYWORDS = ("date", "timestamp")

def is_string_type(t):
    return any(t.lower().startswith(s) for s in STRING_TYPES)

def is_date_named(name):
    n = name.lower()
    return any(kw in n for kw in DATE_KEYWORDS)

# ── styles ─────────────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1F3864")
HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
PK_FILL    = PatternFill("solid", fgColor="FFF2CC")
TBL_FILL   = PatternFill("solid", fgColor="D9E1F2")
ALT_FILL   = PatternFill("solid", fgColor="EEF2F8")
DIFF_FILL  = PatternFill("solid", fgColor="FCE4D6")  # type mismatch src vs tgt
STR_FILL   = PatternFill("solid", fgColor="D9B3FF")  # purple: date-named but string type


def write_sheet(wb, db, src_schema, tgt_schema):
    ws = wb.create_sheet(title=db[:31])  # sheet name max 31 chars

    headers = ["Database", "Table", "Field", "Source Type (MySQL4)", "Target Type (MySQL8)", "Is PK", "PK Order"]
    ws.append(headers)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Union of all tables from both sides
    all_tables = sorted(set(list(src_schema.keys()) + list(tgt_schema.keys())))

    row_idx = 2
    prev_tbl = None
    tbl_fill = TBL_FILL

    for tbl in all_tables:
        src_fields = src_schema.get(tbl, {})
        tgt_fields = tgt_schema.get(tbl, {})
        all_fields = list(dict.fromkeys(list(src_fields.keys()) + list(tgt_fields.keys())))  # ordered union

        if prev_tbl != tbl:
            tbl_fill = TBL_FILL if tbl_fill == ALT_FILL else ALT_FILL
            prev_tbl = tbl

        for field in all_fields:
            src_info = src_fields.get(field, {})
            tgt_info = tgt_fields.get(field, {})
            src_type = src_info.get("type", "—")
            tgt_type = tgt_info.get("type", "—")

            # Include if: type is date/time OR name contains date/timestamp keyword
            DATE_TYPES = ("timestamp", "datetime", "date")
            src_is_date = any(dt in src_type.lower() for dt in DATE_TYPES)
            tgt_is_date = any(dt in tgt_type.lower() for dt in DATE_TYPES)
            src_is_str  = src_type != "—" and is_string_type(src_type)
            tgt_is_str  = tgt_type != "—" and is_string_type(tgt_type)
            date_named_string = is_date_named(field) and (src_is_str or tgt_is_str)

            if not src_is_date and not tgt_is_date and not date_named_string:
                continue

            pk_order  = src_info.get("pk_order") or tgt_info.get("pk_order") or 0
            is_pk     = pk_order > 0

            ws.append([
                db, tbl, field,
                src_type, tgt_type,
                "✓ PK" if is_pk else "",
                pk_order if is_pk else "",
            ])

            row = ws[row_idx]
            fill = PK_FILL if is_pk else tbl_fill
            type_mismatch = (src_type != tgt_type and src_type != "—" and tgt_type != "—")

            fname_lower = field.lower()
            if date_named_string:
                name_fill = STR_FILL   # purple — overrides all name-based colors
            elif fname_lower == "timestamp":
                name_fill = PatternFill("solid", fgColor="C6EFCE")  # green
            elif fname_lower == "date":
                name_fill = PatternFill("solid", fgColor="FFC7CE")  # red
            else:
                name_fill = None

            for ci, cell in enumerate(row, 1):
                cell.fill = fill
                cell.alignment = Alignment(vertical="center")
                if type_mismatch and ci in (4, 5):
                    cell.fill = DIFF_FILL
                if name_fill and ci == 3:  # column 3 = Field
                    cell.fill = name_fill
                if is_pk:
                    cell.font = Font(bold=True, size=10)

            row_idx += 1

    # Column widths
    col_widths = [12, 35, 35, 28, 28, 8, 8]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{row_idx - 1}"


# ── main ───────────────────────────────────────────────────────────────────────
def main():
    print("Fetching database list from target...")
    databases = get_databases_tgt()
    print(f"Found {len(databases)} databases: {databases}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for db in databases:
        print(f"\n[{db}] Loading source schema...", end=" ", flush=True)
        try:
            src_schema = get_schema_src(db)
            print(f"{len(src_schema)} tables", end=" | ", flush=True)
        except Exception as e:
            print(f"ERROR: {e}", end=" | ", flush=True)
            src_schema = {}

        print("Loading target schema...", end=" ", flush=True)
        try:
            tgt_schema = get_schema_tgt(db)
            print(f"{len(tgt_schema)} tables", flush=True)
        except Exception as e:
            print(f"ERROR: {e}", flush=True)
            tgt_schema = {}

        if not src_schema and not tgt_schema:
            print(f"  Skipping {db} — no schema from either side")
            continue

        write_sheet(wb, db, src_schema, tgt_schema)
        print(f"  Sheet '{db}' written")

    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "schema_dump.xlsx")
    wb.save(out_path)
    print(f"\nDone → {out_path}")


if __name__ == "__main__":
    main()
